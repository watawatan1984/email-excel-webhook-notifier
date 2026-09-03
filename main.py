#!/usr/local/bin/python3
# -*- coding: utf-8 -*-

import configparser
import datetime
import json
import logging
import os
import re
import sys
import time
import traceback
import uuid
import io
import unicodedata
import urllib.parse
import base64

# 外部ライブラリ
import openpyxl
import requests
import msoffcrypto
import msal
from requests.auth import HTTPBasicAuth

# --- グローバル設定 ---
CONFIG_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.ini')
MAX_RETRY_COUNT = 3
RETRY_DELAY_SECONDS = 60

# --- 定数 ---
TARGET_CITY = "〇〇市"

# エリア判定用データ構造
TARGET_AREAS = {
    "エリアA": ["町名A1", "町名A2", "町名A3"],
    "エリアB": ["町名B1", "町名B2", "町名B3"],
    "エリアC": ["町名C1", "町名C2", "町名C3"]
}



# --- ロギング設定 ---
def setup_logging(log_dir):
    if not os.path.exists(log_dir):
        os.makedirs(log_dir, exist_ok=True)
    log_file_path = os.path.join(log_dir, 'app.log')

    # 既存のハンドラをクリアして重複ロギングを防ぐ
    logger = logging.getLogger()
    if logger.hasHandlers():
        logger.handlers.clear()

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(module)s - %(funcName)s - %(lineno)d - %(message)s',
        handlers=[
            logging.FileHandler(log_file_path, encoding='utf-8'),
            logging.StreamHandler(sys.stdout) # コンソールにも出力
        ]
    )

# --- 設定読み込み ---
def load_config(config_path):
    if not os.path.exists(config_path):
        sys.stderr.write(f"設定ファイルが見つかりません: {config_path}\n")
        logging.error(f"設定ファイルが見つかりません: {config_path}")
        sys.exit(1)
    config = configparser.ConfigParser(interpolation=None) # interpolation=None は % をエスケープしないようにするため
    config.read(config_path, encoding='utf-8')
    return config

# --- ヘルパー関数 ---
def clean_string(text_val, preserve_internal_spaces_and_dots=False):
    if text_val is None:
        return None
    if not isinstance(text_val, str):
        return str(text_val) # 数値なども文字列として処理

    # NFKC正規化: 全角英数字記号を半角に、半角カタカナを全角に統一など
    cleaned_text = unicodedata.normalize('NFKC', text_val)

    # 不可視文字の除去 (例: ゼロ幅スペース、ソフトハイフン、BOM)
    invisible_chars = ['\u200b', '\u00ad', '\ufeff']
    for char in invisible_chars:
        cleaned_text = cleaned_text.replace(char, '')

    # スペースの処理: preserve_internal_spaces_and_dots が False の場合、連続する空白を1つに
    if not preserve_internal_spaces_and_dots:
        cleaned_text = re.sub(r'\s+', ' ', cleaned_text)

    return cleaned_text.strip() # 前後の空白を除去

def format_excel_date(excel_date_value):
    if excel_date_value is None or str(excel_date_value).strip() == "":
        return ""

    if isinstance(excel_date_value, (int, float)):
        try:
            # Excelの1900年うるう年バグの考慮
            if excel_date_value > 60: # 1900/2/29 (Excelのシリアル値60) 以降の日付
                 excel_date_value -=1 # 補正
            base_date = datetime.datetime(1899, 12, 31) # 0日目
            dt = base_date + datetime.timedelta(days=excel_date_value)
            return dt.strftime('%Y年%m月%d日')
        except (ValueError, TypeError) as e:
            logging.debug(f"Could not convert numeric date '{excel_date_value}'. Error: {e}. Returning as string.")
            return str(excel_date_value)
    elif isinstance(excel_date_value, datetime.datetime):
        return excel_date_value.strftime('%Y年%m月%d日')
    elif isinstance(excel_date_value, str):
        date_str_cleaned = clean_string(excel_date_value)
        if not date_str_cleaned: return ""

        # すでに "YYYY年MM月DD日" 形式ならそのまま返す
        if '年' in date_str_cleaned and '月' in date_str_cleaned and '日' in date_str_cleaned:
            return date_str_cleaned

        # 'YYYY/MM/DD' or 'YYYY-MM-DD' (+時刻情報) のような形式を試す
        date_part = date_str_cleaned.split()[0] # 時刻情報があれば除去

        for fmt in ('%Y/%m/%d', '%Y-%m-%d'):
            try:
                dt = datetime.datetime.strptime(date_part, fmt)
                return dt.strftime('%Y年%m月%d日')
            except ValueError:
                continue
        # 数値文字列かもしれないので再帰的に試す
        try:
            num_val = float(date_part)
            return format_excel_date(num_val) # 数値として再処理
        except ValueError:
            pass # 数値でなければ次の処理へ

        logging.debug(f"Could not parse string date '{date_str_cleaned}'. Returning original.")
        return date_str_cleaned # 解析できなければ元の文字列

    logging.debug(f"Unknown date type for '{excel_date_value}' ({type(excel_date_value)}). Returning as string.")
    return str(excel_date_value)

def get_cell_value(sheet, row_idx, col_idx, preserve_internal_spaces_and_dots=False):
    try:
        cell = sheet.cell(row=row_idx, column=col_idx)
        raw_value = None

        # 結合セルチェック: 指定されたセルが結合範囲内か確認し、そうであれば左上のセルの値を取得
        for merged_range in sheet.merged_cells.ranges:
            if merged_range.min_row <= cell.row <= merged_range.max_row and \
               merged_range.min_col <= cell.column <= merged_range.max_col:
                # 結合セルであれば、範囲の左上のセルの値を取得
                top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                raw_value = top_left_cell.value
                break # 最初の合致で終了

        if raw_value is None: # 結合セルではなかった、または結合範囲の左上でなかった
            raw_value = cell.value

        return clean_string(raw_value, preserve_internal_spaces_and_dots)
    except Exception as e:
        logging.error(f"Error getting cell value at ({row_idx}, {col_idx}): {e}")
        logging.debug(traceback.format_exc())
        return None

def get_combined_value(sheet, row_idx, start_col_idx, end_col_idx, preserve_internal_spaces_and_dots_for_all=False):
    values = []
    for c_idx in range(start_col_idx, end_col_idx + 1):
        cell_val = get_cell_value(sheet, row_idx, c_idx,
                                  preserve_internal_spaces_and_dots=preserve_internal_spaces_and_dots_for_all)
        values.append(str(cell_val or '')) # Noneの場合は空文字に
    return "".join(values).strip() # 結合して前後の空白を除去

def process_phone_number(phone_str_raw):
    if phone_str_raw is None or str(phone_str_raw).strip() == "":
        return "", ""

    text = unicodedata.normalize('NFKC', str(phone_str_raw)).strip()
    main_phone_number = ""
    additional_info_parts = []

    # 電話番号のパターン
    phone_pattern = re.compile(
        r"""\b(
            0\d{1,4}            # 市外局番 (例: 03, 048, 0465)
            (?:[\s-]?\d{1,4}){1,2} # 市内局番と加入者番号の一部 (ハイフンやスペース区切りを許容)
            \d{1,4}             # 加入者番号の残り
            )\b
            |
            \b(0\d{9,10})\b    # ハイフンなしの10桁または11桁の番号 (例: 09012345678)
        """, re.VERBOSE
    )

    # カッコ内の情報を抽出 (例: (本人), (妻), 【携帯】 など)
    owner_pattern = re.compile(r"[\(（\[【]([^)）\]】]+)[\)）\]】]")
    remaining_text_for_phone_search = text # 電話番号検索対象の文字列

    # 先にカッコ内の情報を抽出して除去
    owner_matches = owner_pattern.findall(remaining_text_for_phone_search)
    for owner_info in owner_matches:
        additional_info_parts.append(owner_info.strip())
        remaining_text_for_phone_search = owner_pattern.sub("", remaining_text_for_phone_search, count=1).strip() # 最初のマッチのみ置換して除去

    # 電話番号を抽出
    extracted_phones = []
    # 複数の電話番号が区切り文字で並んでいる場合を考慮
    potential_phone_segments = re.split(r'[/\n、。・]', remaining_text_for_phone_search)

    for segment in potential_phone_segments:
        segment = segment.strip()
        if not segment: continue

        matches = phone_pattern.finditer(segment)
        for match in matches:
            phone_num = next((g for g in match.groups() if g is not None), None)
            if phone_num:
                cleaned_phone = re.sub(r'[^\d]', '', phone_num) # 数字以外を除去
                # 有効な電話番号の桁数チェック (日本国内を想定)
                if cleaned_phone.startswith("0") and 9 <= len(cleaned_phone) <= 11:
                    extracted_phones.append(cleaned_phone)

    if extracted_phones:
        main_phone_number = extracted_phones[0] # 最初の有効な番号をメインとする
        if len(extracted_phones) > 1:
            # 2つ目以降の電話番号を追加情報として記録
            for extra_ph in extracted_phones[1:]:
                additional_info_parts.append(f"追加電話: {extra_ph}")

    final_additional_info = " ".join(filter(None, additional_info_parts)).strip() # 空の要素を除いて結合
    return main_phone_number, final_additional_info

def _format_date_for_mysql(date_str_jp):
    if not date_str_jp:
        return None
    try:
        # 'YYYY年MM月DD日' 形式を 'YYYY-MM-DD' に変換
        dt_obj = datetime.datetime.strptime(date_str_jp, '%Y年%m月%d日')
        return dt_obj.strftime('%Y-%m-%d')
    except ValueError:
        logging.warning(f"日付文字列 '{date_str_jp}' のMySQL用フォーマット変換に失敗しました。")
        return None # 変換失敗時はNoneを返す

def determine_area_name(address_text):
    if not address_text:
        return "不明エリア"

    # 住所文字列の先頭が対象市名であれば、それを取り除いて町名マッチングに使用
    if address_text.startswith(TARGET_CITY):
        address_text_for_match = address_text[len(TARGET_CITY):]
    else:
        address_text_for_match = address_text

    for area, towns in TARGET_AREAS.items():
        for town in towns:
            if town in address_text_for_match: # 町名が住所に含まれていれば該当エリア
                return area
    return "その他エリア" # どのエリアにも合致しなかった場合

# --- Excelデータ抽出 ---
def extract_requests(sheet, request_type):
    requests_data = []
    logging.info(f"Starting extraction for type: {request_type} from sheet: {sheet.title}")

    try:
        # request_type に応じて設定値を決定
        if request_type == "設置":
            first_card_data_start_row = 15 # 最初のカードのデータ開始行
            card_row_interval = 12         # カード間の行間隔
            max_cards_to_check = 8         # チェックするカードの最大数
            col_map = {                    # 各データの列マッピング
                "order_code": (15, 3, 4),  # (行オフセット, 開始列, 終了列) or (行オフセット, 列)
                "request_date": (16, 3),
                "name": (17, 3),
                "phone": (18, 3),
                "address": (19, 3),
                "map_no": (20, 3, 4),
                "preferred_date": (21, 3),
                "remarks": (22, 3),
                "sent_date": (23, 3),
                "request_id": (13, 4)      # 元整理番号 (相対位置は first_card_data_start_row からのオフセットで計算)
            }
        elif request_type == "回収":
            first_card_data_start_row = 13
            card_row_interval = 10
            max_cards_to_check = 12
            col_map = {
                "order_code": (13, 3, 4),
                "name": (14, 3),
                "phone": (15, 3),
                "address": (16, 3),
                "request_date": (17, 3),
                "sent_date": (18, 3),
                "remarks": (19, 3),
                "map_no": None,            # 回収には住宅地図番号なし
                "preferred_date": None,    # 回収には設置希望日なし
                "request_id": (11, 4)
            }
        else:
            logging.error(f"Unknown request type '{request_type}'")
            return []

        for i in range(max_cards_to_check):
            current_row_offset = i * card_row_interval # 現在のカードブロックの行オフセット
            current_block_data = {"依頼分類_raw": request_type} # 抽出データを格納する辞書

            # 氏名を取得 (これが空ならデータ終了とみなす)
            name_abs_row = col_map["name"][0] + current_row_offset
            name_val = get_cell_value(sheet, name_abs_row, col_map["name"][1], preserve_internal_spaces_and_dots=True)
            current_block_data["氏名"] = str(name_val or '').strip()

            if not current_block_data["氏名"]:
                if i == 0: logging.info(f"シート '{sheet.title}' の最初のカード ({request_type}) で氏名が見つかりません。データ無と判断。"); break
                else: logging.info(f"シート '{sheet.title}' のカードブロック {i+1} ({request_type}) で氏名が見つかりません。このブロック以降のデータは無いと判断。"); break

            # 指示書送付日 (これが空ならそのカードはスキップ)
            sent_date_abs_row = col_map["sent_date"][0] + current_row_offset
            sent_date_val = get_cell_value(sheet, sent_date_abs_row, col_map["sent_date"][1])
            current_block_data["指示書送付日"] = format_excel_date(sent_date_val)
            if not current_block_data["指示書送付日"]:
                logging.warning(f"{request_type}カード (ブロック {i+1}, 氏名: {current_block_data['氏名']}) の指示書送付日がありません。このカードをスキップします。")
                continue

            # 元整理番号 (request_id)
            # request_idの行は、カードの基準行 (first_card_data_start_row) からの相対オフセットで計算
            relative_offset_for_req_id = col_map["request_id"][0] - first_card_data_start_row
            req_id_abs_row = first_card_data_start_row + current_row_offset + relative_offset_for_req_id
            req_id_val = get_cell_value(sheet, req_id_abs_row, col_map["request_id"][1])
            current_block_data["元整理番号"] = str(req_id_val or '').strip()

            # 元指示書No (order_code)
            oc_abs_row = col_map["order_code"][0] + current_row_offset
            current_block_data["元指示書No"] = str((get_combined_value(sheet, oc_abs_row, col_map["order_code"][1], col_map["order_code"][2])
                                          if len(col_map["order_code"]) == 3 # 結合セルの場合
                                          else get_cell_value(sheet, oc_abs_row, col_map["order_code"][1])) or '').strip()

            # 電話番号 (phone)
            phone_abs_row = col_map["phone"][0] + current_row_offset
            phone_val_raw = get_cell_value(sheet, phone_abs_row, col_map["phone"][1])
            main_phone, phone_remarks = process_phone_number(phone_val_raw)
            current_block_data["電話番号"] = main_phone

            # 設置場所/住所 (address)
            addr_abs_row = col_map["address"][0] + current_row_offset
            addr_str = str(get_cell_value(sheet, addr_abs_row, col_map["address"][1]) or '').strip()
            # 住所に対象市名が含まれていなければ先頭に付加 (ただし、既に市名がある場合は重複させない)
            current_block_data["設置場所"] = f"{TARGET_CITY}{addr_str}" if addr_str and not addr_str.startswith(TARGET_CITY) and TARGET_CITY not in addr_str else addr_str


            # 元依頼日 (request_date)
            req_date_abs_row = col_map["request_date"][0] + current_row_offset
            current_block_data["元依頼日"] = format_excel_date(get_cell_value(sheet, req_date_abs_row, col_map["request_date"][1]))

            # 元特記事項 (remarks)
            remarks_abs_row = col_map["remarks"][0] + current_row_offset
            remarks_str = str(get_cell_value(sheet, remarks_abs_row, col_map["remarks"][1]) or '').strip()
            if phone_remarks: # 電話番号処理で追加情報があれば備考に追記
                remarks_str = f"{remarks_str} [{phone_remarks}]".strip()
            current_block_data["元特記事項"] = remarks_str

            if request_type == "設置":
                current_block_data["住宅地図"] = str((get_combined_value(sheet, col_map["map_no"][0] + current_row_offset, col_map["map_no"][1], col_map["map_no"][2])
                                             if col_map["map_no"] and len(col_map["map_no"]) == 3
                                             else get_cell_value(sheet, col_map["map_no"][0] + current_row_offset, col_map["map_no"][1]) if col_map["map_no"] else "") or '').strip()
                current_block_data["元設置希望日"] = format_excel_date(get_cell_value(sheet, col_map["preferred_date"][0] + current_row_offset, col_map["preferred_date"][1])) if col_map["preferred_date"] else ""
            else: # 回収の場合
                current_block_data["住宅地図"], current_block_data["元設置希望日"] = "", ""

            current_block_data["Excelシート名"] = sheet.title
            current_block_data["Excel行"] = first_card_data_start_row + current_row_offset # 記録用
            requests_data.append(current_block_data)
            logging.info(f"抽出成功: {request_type} ブロック {i+1}, 氏名: '{current_block_data['氏名']}', 指示書No: '{current_block_data['元指示書No']}', 整理番号: '{current_block_data['元整理番号']}'")

    except Exception as e:
        logging.error(f"Excelデータ抽出中に予期せぬエラーが発生しました ({request_type}, シート: {sheet.title}): {e}\n{traceback.format_exc()}")

    return requests_data

def normalize_shijisho_no(raw_no_val):
    if raw_no_val is None: return None, "0"
    raw_no_str = str(raw_no_val).strip()
    if not raw_no_str: return None, "0"

    match = re.match(r"(\d+)([ｰー-]?\d*F)?([ｰー-]?\d+)?", raw_no_str, re.IGNORECASE)
    if match:
        main_no = match.group(1)
        sub_parts = [p.lstrip('ｰー-') for p in [match.group(2), match.group(3)] if p]
        sub_no = "".join(sub_parts) if sub_parts else "0"
        return main_no, sub_no

    cleaned = re.sub(r'[^\dFf-]', '', raw_no_str)
    parts = cleaned.split('-', 1)
    if len(parts) == 2 and parts[0] and parts[1]:
        return parts[0], parts[1]
    elif parts[0]:
        return parts[0], "0"

    return raw_no_str, "0"

def determine_flags_revised(main_request_type_str_excel, detail_remarks_list_str):
    flags = {
        "フラグ_再設置希望": False,
        "フラグ_捕獲個体": False,
        "フラグ_要事前連絡": False,
        "フラグ_交換希望": False
    }
    combined_remarks_lower = " ".join(str(remark or "") for remark in detail_remarks_list_str).lower()
    main_request_type_lower = main_request_type_str_excel.lower()

    if "再設置希望" in combined_remarks_lower or "わな回収再設置" in main_request_type_lower:
        flags["フラグ_再設置希望"] = True
    if "アライグマ" in combined_remarks_lower or "ハクビシン" in combined_remarks_lower:
        flags["フラグ_捕獲個体"] = True
    if "要事前連絡" in combined_remarks_lower:
        flags["フラグ_要事前連絡"] = True
    if "交換希望" in combined_remarks_lower:
        flags["フラグ_交換希望"] = True

    return flags

def determine_main_work_type(excel_request_type, flags):
    if excel_request_type == "わな設置":
        return "交換設置" if flags.get("フラグ_交換希望") else "わな設置"
    elif excel_request_type == "わな回収":
        return "捕獲回収" if flags.get("フラグ_捕獲個体") else "空きわな回収"
    elif excel_request_type == "わな回収再設置":
        return "捕獲回収再設置" if flags.get("フラグ_捕獲個体") else "個体なし回収再設置"
    return excel_request_type

def create_detail_request_ideal(original_req_data, work_type_str_ideal, detail_type_prefix=""):
    remarks = original_req_data.get("元特記事項", "")
    if detail_type_prefix:
        remarks = f"{detail_type_prefix} {remarks}".strip()

    return {
        "詳細ID": f"{datetime.date.today().strftime('%y%m%d')}-{uuid.uuid4().hex[:8]}",
        "依頼区分": work_type_str_ideal,
        "整理番号": original_req_data.get("元整理番号", ""),
        "関連指示書No": original_req_data.get("元指示書No", ""),
        "指示書送付日": original_req_data.get("元依頼日", ""),
        "住宅地図": original_req_data.get("住宅地図", ""),
        "希望日": original_req_data.get("元設置希望日", ""),
        "特記事項": remarks,
        "status": "対応待ち",
        "memo": ""
    }

def add_single_request_ideal(req_data, raw_work_type_excel, send_date, final_list, config):
    main_req_id_suffix = uuid.uuid4().hex[:8]
    main_req_id = f"{datetime.date.today().strftime('%Y%m%d')}-{main_req_id_suffix}"

    excel_ideal_type = "わな" + raw_work_type_excel
    flags = determine_flags_revised(excel_ideal_type, [req_data.get("元特記事項")])
    main_work_types_value = determine_main_work_type(excel_ideal_type, flags)

    address = req_data.get("設置場所", "")
    area_name = determine_area_name(address)

    map_base_cfg_key = ('API', 'MAP_URL_BASE')
    map_base = config.get(*map_base_cfg_key, fallback='https://www.google.com/maps/search/?api=1&query=') \
        if config.has_section(map_base_cfg_key[0]) and config.has_option(*map_base_cfg_key) \
        else 'https://www.google.com/maps/search/?api=1&query='
    map_url = f"{map_base}{urllib.parse.quote_plus(address)}" if address else ""


    final_list.append({
        "依頼メイン": {
            "依頼ID": main_req_id,
            "指示書送付日": send_date,
            "依頼者氏名": req_data.get("氏名"),
            "電話番号": req_data.get("電話番号"),
            "住所": address,
            "district": TARGET_CITY,
            "area_name": area_name,
            "依頼区分": excel_ideal_type,
            "work_types": main_work_types_value,
            "関連指示書No": req_data.get("元指示書No"),
            "マップURL": map_url,
            "総合特記事項": req_data.get("元特記事項"),
            **flags,
            "status": "対応待ち",
            "memo": ""
        },
        "依頼詳細": [create_detail_request_ideal(req_data, excel_ideal_type)]
    })

def process_requests_ideal(extracted_data, config):
    if not extracted_data: return []

    grouped_by_send_date = {}
    for req in extracted_data:
        send_date = req.get("指示書送付日")
        if send_date:
            grouped_by_send_date.setdefault(send_date, []).append(req)

    final_requests_payload = []

    for send_date, requests_on_date in grouped_by_send_date.items():
        requests_with_norm_no = []
        for r in requests_on_date:
            main_no, _ = normalize_shijisho_no(r.get("元指示書No"))
            group_key = main_no if main_no else f"__ORPHAN_{uuid.uuid4().hex}"
            requests_with_norm_no.append({"data": r, "group_key": group_key})

        grouped_by_internal_key = {}
        for r_norm in requests_with_norm_no:
            grouped_by_internal_key.setdefault(r_norm["group_key"], []).append(r_norm["data"])

        processed_tuples = set()

        for _, group_items_data in grouped_by_internal_key.items():
            installs = [item for item in group_items_data if item.get("依頼分類_raw") == "設置"]
            retrievals = [item for item in group_items_data if item.get("依頼分類_raw") == "回収"]

            available_installs = list(installs)

            for ret_item in retrievals:
                ret_tuple = (ret_item.get("元指示書No"), ret_item.get("氏名"), ret_item.get("依頼分類_raw"))
                if ret_tuple in processed_tuples: continue

                match_install = next((inst_item for inst_item in available_installs
                                      if inst_item.get("氏名") == ret_item.get("氏名") and
                                      (inst_item.get("元指示書No"), inst_item.get("氏名"), inst_item.get("依頼分類_raw")) not in processed_tuples),
                                     None)

                if match_install:
                    main_id_suffix = uuid.uuid4().hex[:8]
                    main_id = f"{datetime.date.today().strftime('%Y%m%d')}-{main_id_suffix}"

                    rem_ret = str(ret_item.get("元特記事項") or "")
                    rem_inst = str(match_install.get("元特記事項") or "")
                    main_remarks = f"回収特記: {rem_ret}\n設置特記: {rem_inst}".strip() if rem_ret or rem_inst else ""

                    flags = determine_flags_revised("わな回収再設置", [rem_ret, rem_inst])
                    main_work_types_value = determine_main_work_type("わな回収再設置", flags)

                    addr = match_install.get("設置場所", "") or ret_item.get("設置場所", "")
                    area_name = determine_area_name(addr)
                    
                    map_base_cfg_key = ('API', 'MAP_URL_BASE')
                    map_base = config.get(*map_base_cfg_key, fallback='https://www.google.com/maps/search/?api=1&query=') \
                        if config.has_section(map_base_cfg_key[0]) and config.has_option(*map_base_cfg_key) \
                        else 'https://www.google.com/maps/search/?api=1&query='
                    map_url = f"{map_base}{urllib.parse.quote_plus(addr)}" if addr else ""

                    final_requests_payload.append({
                        "依頼メイン": {
                            "依頼ID": main_id, "指示書送付日": send_date,
                            "依頼者氏名": ret_item.get("氏名"),
                            "電話番号": ret_item.get("電話番号") or match_install.get("電話番号"),
                            "住所": addr,
                            "district": TARGET_CITY,
                            "area_name": area_name,
                            "依頼区分": "わな回収再設置",
                            "work_types": main_work_types_value,
                            "関連指示書No": ret_item.get("元指示書No") or match_install.get("元指示書No"),
                            "マップURL": map_url,
                            "総合特記事項": main_remarks, **flags,
                            "status": "対応待ち", "memo": ""
                        },
                        "依頼詳細": [
                            create_detail_request_ideal(ret_item, "わな回収", detail_type_prefix="[回収作業]"),
                            create_detail_request_ideal(match_install, "わな設置", detail_type_prefix="[再設置作業]")
                        ]
                    })
                    processed_tuples.add(ret_tuple)
                    processed_tuples.add((match_install.get("元指示書No"), match_install.get("氏名"), match_install.get("依頼分類_raw")))
                    if match_install in available_installs: available_installs.remove(match_install)

            for remainder_list, raw_type_excel in [
                (available_installs, "設置"),
                ([r for r in retrievals if (r.get("元指示書No"), r.get("氏名"), r.get("依頼分類_raw")) not in processed_tuples], "回収")
            ]:
                for item_data in remainder_list:
                    item_tuple = (item_data.get("元指示書No"), item_data.get("氏名"), item_data.get("依頼分類_raw"))
                    if item_tuple not in processed_tuples:
                        add_single_request_ideal(item_data, raw_type_excel, send_date, final_requests_payload, config)
                        processed_tuples.add(item_tuple)

    return final_requests_payload

def extract_master_excel_data(workbook_path, config):
    all_extracted = []
    try:
        workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    except Exception as e:
        logging.error(f"Excelファイルの読み込みに失敗しました ({workbook_path}): {e}\n{traceback.format_exc()}")
        return []

    logging.info(f"Excelファイル読み込み成功: {workbook_path}. 利用可能なシート: {workbook.sheetnames}")

    for sheet_name_key, req_type in [
        ('SHEET_NAME_INSTALL', "設置"),
        ('SHEET_NAME_RECOVERY', "回収")
    ]:
        sheet_name = config.get('Excel', sheet_name_key, fallback=None)
        if sheet_name and sheet_name in workbook.sheetnames:
            all_extracted.extend(extract_requests(workbook[sheet_name], req_type))
        elif sheet_name:
            logging.warning(f"設定されたシート名 '{sheet_name}' (キー: {sheet_name_key}) がExcelファイル内に見つかりません。")
        else:
            logging.warning(f"シート名がconfigファイルで設定されていません (キー: {sheet_name_key})。")

    return all_extracted

# --- メール処理 (Microsoft Graph API版) ---

def get_graph_access_token(config):
    """Azure ADからGraph APIのアクセストークンを取得"""
    try:
        client_id = config['Email']['CLIENT_ID']
        client_secret = config['Email']['CLIENT_SECRET']
        tenant_id = config['Email']['TENANT_ID']
        
        authority = f"https://login.microsoftonline.com/{tenant_id}"
        app = msal.ConfidentialClientApplication(
            client_id, authority=authority, client_credential=client_secret
        )
        # バックグラウンドアプリ用のスコープ
        result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
        
        if "access_token" in result:
            logging.info("Graph API アクセストークンの取得に成功しました。")
            return result["access_token"]
        else:
            logging.error(f"トークン取得エラー: {result.get('error')} - {result.get('error_description')}")
            return None
    except Exception as e:
        logging.error(f"トークン取得中に例外が発生しました: {e}\n{traceback.format_exc()}")
        return None

def search_emails_graph(token, config, processed_ids_path):
    """Graph APIでメールを検索 (差出人指定、添付あり)"""
    processed_ids = load_processed_email_uids(processed_ids_path)
    target_email = config['Email']['TARGET_EMAIL_ADDRESS']
    target_sender = config['Email']['TARGET_SENDER']
    
    endpoint = f"https://graph.microsoft.com/v1.0/users/{target_email}/messages"
    headers = {'Authorization': 'Bearer ' + token}
    
    # フィルタ: 指定差出人 かつ 添付あり かつ 未読
    # ※API側で未読(isRead eq false)のみに絞ることで効率化
    query = f"$filter=from/emailAddress/address eq '{target_sender}' and hasAttachments eq true and isRead eq false"
    # 必要なフィールドのみ取得
    query += "&$select=id,subject,hasAttachments,receivedDateTime"
    # 最新10件程度取得
    query += "&$top=10"
    
    logging.info(f"Graph API メール検索クエリ: {query}")
    
    try:
        response = requests.get(f"{endpoint}?{query}", headers=headers, timeout=30)
        response.raise_for_status()
        data = response.json()
        messages = data.get('value', [])
        
        new_emails = []
        for msg in messages:
            msg_id = msg['id']
            subject = msg.get('subject', 'No Subject')
            
            if msg_id in processed_ids:
                logging.debug(f"メッセージID {msg_id[-10:]}... は既に処理済みです。")
                continue
            
            logging.info(f"処理対象のメールを発見: ID={msg_id[-10:]}..., 件名='{subject}'")
            new_emails.append({'id': msg_id, 'subject': subject})
            
        return new_emails

    except Exception as e:
        logging.error(f"メール検索中にエラーが発生しました: {e}\n{traceback.format_exc()}")
        return []

def download_excel_attachment_graph(token, config, msg_id, temp_dir_path):
    """Graph APIで添付ファイルを取得しExcelを保存"""
    target_email = config['Email']['TARGET_EMAIL_ADDRESS']
    endpoint = f"https://graph.microsoft.com/v1.0/users/{target_email}/messages/{msg_id}/attachments"
    headers = {'Authorization': 'Bearer ' + token}
    
    try:
        response = requests.get(endpoint, headers=headers, timeout=60)
        response.raise_for_status()
        attachments = response.json().get('value', [])
        
        for att in attachments:
            filename = att.get('name', '')
            # Excelファイルか確認
            if filename.lower().endswith(('.xlsx', '.xlsm')):
                content_bytes = att.get('contentBytes')
                if not content_bytes:
                    continue
                
                # Base64デコード
                file_content = base64.b64decode(content_bytes)
                
                # ファイル名生成
                safe_fn = re.sub(r'[\\/*?:"<>|]', "_", filename)
                unique_fn = f"{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:8]}_{safe_fn}"
                filepath = os.path.join(temp_dir_path, unique_fn)
                
                if not os.path.exists(temp_dir_path):
                    os.makedirs(temp_dir_path, exist_ok=True)
                
                with open(filepath, 'wb') as f:
                    f.write(file_content)
                
                logging.info(f"Excel添付ファイルのダウンロードに成功: {filepath}")
                return filepath # 最初のExcelを見つけたら返す
        
        logging.warning("メール内にExcel添付ファイルが見つかりませんでした。")
        return None

    except Exception as e:
        logging.error(f"添付ファイル取得エラー (MsgID: {msg_id[-10:]}...): {e}\n{traceback.format_exc()}")
        return None

def mark_email_as_processed_graph(token, config, msg_id, processed_ids_filepath):
    """Graph APIでメールを既読にする & ローカル記録"""
    target_email = config['Email']['TARGET_EMAIL_ADDRESS']
    endpoint = f"https://graph.microsoft.com/v1.0/users/{target_email}/messages/{msg_id}"
    headers = {
        'Authorization': 'Bearer ' + token,
        'Content-Type': 'application/json'
    }
    payload = {'isRead': True}
    
    try:
        response = requests.patch(endpoint, headers=headers, json=payload, timeout=30)
        response.raise_for_status()
        logging.info(f"メールを既読に設定しました (ID: {msg_id[-10:]}...)")
        
        # IDをローカルファイルに記録
        add_processed_email_uid(msg_id, processed_ids_filepath)
        
    except Exception as e:
        logging.error(f"既読設定エラー (MsgID: {msg_id[-10:]}...): {e}")

def decrypt_excel(encrypted_path, password_from_config):
    base, ext = os.path.splitext(encrypted_path)
    decrypted_path = f"{base}_decrypted{ext}" # 復号後のファイルパス

    try:
        with open(encrypted_path, "rb") as f_enc:
            office_file = msoffcrypto.OfficeFile(f_enc)
            office_file.load_key(password=password_from_config) # パスワードでキーをロード
            with open(decrypted_path, "wb") as f_dec:
                office_file.decrypt(f_dec) # 復号して書き込み
        logging.info(f"Excelファイルの復号に成功しました: {decrypted_path}")
        return decrypted_path
    except Exception as e:
        logging.error(f"Excelファイルの復号に失敗しました ({encrypted_path}): {e}\n{traceback.format_exc()}")
        if os.path.exists(decrypted_path):
            try: os.remove(decrypted_path)
            except: pass
        return None

def load_processed_email_uids(filepath):
    if not os.path.exists(filepath):
        return set() # ファイルがなければ空のセット
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return set(line.strip() for line in f if line.strip())
    except Exception as e:
        logging.error(f"処理済みIDファイルの読み込みに失敗しました ({filepath}): {e}")
        return set()

def add_processed_email_uid(uid, filepath):
    try:
        with open(filepath, 'a', encoding='utf-8') as f:
            f.write(uid + '\n')
    except Exception as e:
        logging.error(f"処理済みIDのファイル記録に失敗しました ({filepath}): {e}")

# --- WordPress REST API 連携 ---

def get_wp_base_url(config):
    base_url = config.get('API', 'WP_BASE_URL', fallback='').strip()
    if not base_url: return ''
    return base_url[:-1] if base_url.endswith('/') else base_url

def get_wp_auth(config):
    username = config.get('API', 'WP_APP_USERNAME', fallback='').strip()
    password = config.get('API', 'WP_APP_PASSWORD', fallback='')
    if not username or not password:
        logging.error("WP REST APIの認証情報が設定されていません。")
        return None
    return HTTPBasicAuth(username, password)

def bool_to_int(value):
    if isinstance(value, bool): return 1 if value else 0
    if isinstance(value, (int, float)): return 1 if value else 0
    if isinstance(value, str): return 1 if value.strip().lower() in ('1', 'true', 'yes', 'on') else 0
    return 0

def build_wp_request_payload(processed_item, config):
    main_data = processed_item.get("依頼メイン", {})
    if not main_data: return None

    request_id = main_data.get("依頼ID") or f"AUTO-{uuid.uuid4().hex}"
    sent_date_iso = _format_date_for_mysql(main_data.get("指示書送付日")) or datetime.date.today().isoformat()
    address = main_data.get("住所") or ''
    map_url = main_data.get("マップURL")

    if (not map_url) and address:
        map_base_key = ('API', 'MAP_URL_BASE')
        map_base = config.get(*map_base_key, fallback='https://www.google.com/maps/search/?api=1&query=') \
            if config.has_section(map_base_key[0]) and config.has_option(*map_base_key) \
            else 'https://www.google.com/maps/search/?api=1&query='
        map_url = f"{map_base}{urllib.parse.quote_plus(address)}"

    payload = {
        'request_id': request_id,
        'sent_date': sent_date_iso,
        'client_name': main_data.get("依頼者氏名"),
        'phone_number': main_data.get("電話番号"),
        'address': address,
        'district': main_data.get("district") or TARGET_CITY,
        'area_name': main_data.get("area_name") or determine_area_name(address),
        'request_type': main_data.get("依頼区分"),
        'work_types': main_data.get("work_types"),
        'related_instruction_no': main_data.get("関連指示書No"),
        'map_url': map_url,
        'remarks_main': main_data.get("総合特記事項"),
        'flag_reinstall_request': bool_to_int(main_data.get("フラグ_再設置希望", False)),
        'flag_captured_animal': bool_to_int(main_data.get("フラグ_捕獲個体", False)),
        'flag_needs_prior_contact': bool_to_int(main_data.get("フラグ_要事前連絡", False)),
        'flag_needs_trade': bool_to_int(main_data.get("フラグ_交換希望", False)),
        'status': main_data.get("status", "対応待ち"),
        'memo': main_data.get("memo", "")
    }
    return payload

def send_requests_via_rest(processed_data_list, config):
    if not processed_data_list:
        logging.info("REST APIへの送信対象データはありません。")
        return True, 0, None, None

    base_url = get_wp_base_url(config)
    endpoint_path = config.get('API', 'WP_BULK_REQUESTS_ENDPOINT', fallback='').strip()
    if not base_url or not endpoint_path:
        logging.error("WP REST APIのエンドポイント情報が不足しています。")
        return False, 0, None, None

    auth = get_wp_auth(config)
    if auth is None: return False, 0, None, None

    url = f"{base_url}{endpoint_path}"
    payloads = []
    for item in processed_data_list:
        p = build_wp_request_payload(item, config)
        if p: payloads.append(p)

    if not payloads:
        logging.warning("生成されたRESTペイロードが空のため送信をスキップします。")
        return True, 0, None, None

    request_body = {'requests': payloads}
    headers = {'Content-Type': 'application/json'}

    try:
        response = requests.post(url, auth=auth, headers=headers, json=request_body, timeout=30)
        try:
            response_json = response.json()
        except ValueError:
            response_json = {'raw': response.text}
    except requests.exceptions.RequestException as e_req:
        logging.error(f"WP REST APIへの送信中にリクエストエラーが発生しました: {e_req}")
        return False, 0, None, None

    status_code = response.status_code

    if 200 <= status_code < 300:
        created_info = response_json.get('created_info')
        if isinstance(created_info, list):
            created_count = len(created_info)
        else:
            data_array = response_json.get('data')
            created_count = len(data_array) if isinstance(data_array, list) else len(payloads)
        logging.info(f"WP REST APIへの送信に成功しました。登録件数: {created_count}")
        return True, created_count, response_json, status_code

    if status_code == 409:
        logging.warning(f"WP REST APIが重複を検出しました: {response_json}")
        return True, 0, response_json, status_code

    logging.error(f"WP REST APIへの送信が失敗しました (HTTP {status_code}): {response_json}")
    return False, 0, response_json, status_code

# --- 通知関連 ---
def generate_teams_message(processed_data_list):
    if not processed_data_list: return None

    counts = {
        "わな設置": 0, "交換設置": 0, "空きわな回収": 0, "捕獲回収": 0,
        "個体なし回収再設置": 0, "捕獲回収再設置": 0,
        "（元）わな設置": 0, "（元）わな回収":0, "（元）わな回収再設置":0
    }

    for req_obj in processed_data_list:
        main_req = req_obj.get("依頼メイン", {})
        work_type = main_req.get("work_types", "")
        original_type = main_req.get("依頼区分", "")

        if work_type and work_type in counts: counts[work_type] += 1
        if original_type == "わな設置": counts["（元）わな設置"] +=1
        if original_type == "わな回収": counts["（元）わな回収"] +=1
        if original_type == "わな回収再設置": counts["（元）わな回収再設置"] +=1

    msg_parts = []
    if counts["交換設置"] > 0: msg_parts.append(f"交換設置({counts['交換設置']}件)")
    elif counts["わな設置"] > 0: msg_parts.append(f"新規設置({counts['わな設置']}件)")

    if counts["捕獲回収再設置"] > 0: msg_parts.append(f"捕獲回収再設置({counts['捕獲回収再設置']}件)")
    elif counts["個体なし回収再設置"] > 0: msg_parts.append(f"空わな回収再設置({counts['個体なし回収再設置']}件)")

    if counts["捕獲回収"] > 0: msg_parts.append(f"捕獲回収({counts['捕獲回収']}件)")
    elif counts["空きわな回収"] > 0: msg_parts.append(f"空わな回収({counts['空きわな回収']}件)")

    if not msg_parts:
        if counts["（元）わな設置"] > 0: msg_parts.append(f"設置(元分類)({counts['（元）わな設置']}件)")
        if counts["（元）わな回収"] > 0: msg_parts.append(f"回収(元分類)({counts['（元）わな回収']}件)")
        if counts["（元）わな回収再設置"] > 0: msg_parts.append(f"回収再設置(元分類)({counts['（元）わな回収再設置']}件)")

    if not msg_parts:
        return f"{TARGET_CITY}のわな依頼がありましたが、詳細は不明です（登録件数{len(processed_data_list)}）。"

    return f"{TARGET_CITY}からわなの「{'」と「'.join(msg_parts)}」依頼が届きました。"

def send_teams_notification(message_text, config):
    if not message_text: return False
    if not config.has_section('Teams') or not config['Teams'].get('WEBHOOK_URL'): return False

    webhook_url = config['Teams']['WEBHOOK_URL']
    dashboard_url = config.get('Teams', 'DASHBOARD_URL', fallback='')

    adaptive_card_content = {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.5",
        "body": [{"type": "TextBlock", "text": message_text, "wrap": True, "size": "Large"}]
    }
    if dashboard_url:
        adaptive_card_content["actions"] = [{"type": "Action.OpenUrl", "title": "ダッシュボードで確認", "url": dashboard_url}]

    teams_payload_structure = config.get('Teams', 'PAYLOAD_STRUCTURE', fallback='gas_logic_apps_style').lower()
    payload_to_send = {}

    if teams_payload_structure == 'direct_adaptive_card':
         payload_to_send = {
            "type": "message",
            "attachments": [{"contentType": "application/vnd.microsoft.card.adaptive", "content": adaptive_card_content}]
        }
    else: # gas_logic_apps_style
        payload_to_send = { "body": { "attachments": [{"contentType": "application/vnd.microsoft.card.adaptive", "contentUrl": None, "content": adaptive_card_content}]}}

    try:
        response = requests.post(webhook_url, headers={"Content-Type": "application/json"}, data=json.dumps(payload_to_send), timeout=15)
        if response.status_code in [200, 202] or response.text == "1":
            logging.info(f"Teams通知成功: {response.status_code}")
            return True
        else:
            logging.error(f"Teams通知エラー: {response.status_code} - {response.text}")
            return False
    except Exception as e:
        logging.error(f"Teams通知例外: {e}")
        return False

def send_onesignal_notification(title, message_body, badge_count, config):
    if not config.has_section('OneSignal'): return False
    app_id = config.get('OneSignal', 'APP_ID')
    api_key = config.get('OneSignal', 'API_KEY')
    if not app_id or not api_key: return False

    payload = {
        "app_id": app_id, "included_segments": ["Active Subscriptions"],
        "headings": {"en": title, "ja": title}, "contents": {"en": message_body, "ja": message_body},
        "ios_badgeType": "SetTo", "ios_badgeCount": badge_count,
        "android_badgeCount": badge_count, "chrome_web_badge": str(badge_count) if badge_count > 0 else "",
        "data": {"app_badge_count": badge_count, "notification_title": title, "notification_body": message_body}
    }
    target_url = config.get('OneSignal', 'TARGET_URL', fallback=config.get('Teams', 'DASHBOARD_URL'))
    if target_url: payload["web_url"] = target_url

    try:
        requests.post("https://onesignal.com/api/v1/notifications", headers={"Content-Type": "application/json", "Authorization": f"Basic {api_key}"}, data=json.dumps(payload), timeout=15)
        logging.info("OneSignal通知成功")
        return True
    except Exception as e:
        logging.error(f"OneSignal通知例外: {e}")
        return False

# --- メイン処理 ---
def main():
    config = load_config(CONFIG_FILE_PATH)
    setup_logging(config.get('Files', 'LOG_DIR', fallback=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')))
    logging.info("メール処理およびわな依頼自動登録スクリプトを開始します (Microsoft Graph API版)")

    temp_dir = config.get('Files', 'TEMP_DIR', fallback=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'temp'))
    processed_emails_file = config.get('Files', 'PROCESSED_EMAILS_FILE', fallback=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'processed_uids.txt'))
    if not os.path.exists(temp_dir): os.makedirs(temp_dir, exist_ok=True)

    if not config.has_section('API'):
        logging.critical("API設定が見つかりません。終了します。")
        sys.exit(1)

    # 1. Graph API トークン取得
    access_token = get_graph_access_token(config)
    if not access_token:
        logging.critical("Graph APIトークン取得失敗。終了します。")
        sys.exit(1)

    successfully_processed_payloads_for_teams = []
    total_newly_inserted_main_requests_count = 0
    email_processing_attempted = False
    teams_notification_messages = []

    try:
        # 2. メール検索 (Graph API)
        new_email_list = search_emails_graph(access_token, config, processed_emails_file)

        if not new_email_list:
            logging.info("新規メールはありませんでした。")
        else:
            email_processing_attempted = True
            logging.info(f"{len(new_email_list)}件の新規メールが見つかりました。")

            for email_info in new_email_list:
                msg_id = email_info['id']
                logging.info(f"メール処理開始 (ID: {msg_id[-10:]}...)")

                # 3. 添付ファイル取得 (Graph API)
                enc_excel_path = download_excel_attachment_graph(access_token, config, msg_id, temp_dir)

                if not enc_excel_path:
                    logging.info(f"Excel添付なし (ID: {msg_id[-10:]}...)。処理済みとしてマークします。")
                    mark_email_as_processed_graph(access_token, config, msg_id, processed_emails_file)
                    continue

                # 4. 復号処理
                dec_excel_path = decrypt_excel(enc_excel_path, config['Excel']['PASSWORD'])
                
                # 元ファイル削除
                if os.path.exists(enc_excel_path):
                    try: os.remove(enc_excel_path)
                    except: pass

                if not dec_excel_path:
                    logging.error(f"復号失敗 (ID: {msg_id[-10:]}...)。処理済みとしてマークしスキップします。")
                    mark_email_as_processed_graph(access_token, config, msg_id, processed_emails_file)
                    continue

                # 5. データ抽出
                extracted_raw_data = extract_master_excel_data(dec_excel_path, config)

                # 復号済みファイル削除
                if os.path.exists(dec_excel_path):
                    try: os.remove(dec_excel_path)
                    except: pass

                if not extracted_raw_data:
                    logging.info(f"有効なデータなし (ID: {msg_id[-10:]}...)。処理済みとします。")
                    mark_email_as_processed_graph(access_token, config, msg_id, processed_emails_file)
                    continue

                # 6. ペイロード生成 & API送信
                db_ready_payloads = process_requests_ideal(extracted_raw_data, config)
                if db_ready_payloads:
                    api_success, inserted_count, api_response, http_status = send_requests_via_rest(db_ready_payloads, config)

                    if api_success:
                        logging.info(f"API送信成功: {inserted_count}件 (ID: {msg_id[-10:]}...)")
                        if inserted_count > 0:
                             total_newly_inserted_main_requests_count += inserted_count
                        
                        # 7. メールを既読にする
                        mark_email_as_processed_graph(access_token, config, msg_id, processed_emails_file)
                        
                        successfully_processed_payloads_for_teams.extend(db_ready_payloads)
                        if http_status == 409: logging.info("重複スキップあり")
                    else:
                        logging.error(f"API送信失敗 (ID: {msg_id[-10:]}...): {api_response}")
                        teams_notification_messages.append(
                            f"**API登録失敗**: メールID `{msg_id[-10:]}...` の保存失敗。"
                        )
                else:
                    logging.info(f"ペイロード生成なし (ID: {msg_id[-10:]}...)。処理済みとします。")
                    mark_email_as_processed_graph(access_token, config, msg_id, processed_emails_file)

            # --- 通知処理 ---
            
            # Teams通知メッセージ生成
            if successfully_processed_payloads_for_teams:
                success_summary = generate_teams_message(successfully_processed_payloads_for_teams)
                if success_summary:
                    teams_notification_messages.insert(0, success_summary)

            # Teams送信
            if teams_notification_messages:
                final_teams_message = "\n\n---\n\n".join(teams_notification_messages)
                send_teams_notification(final_teams_message, config)
            elif email_processing_attempted:
                # メール処理は走ったがDB更新なしの場合
                no_db_msg = f"{TARGET_CITY} わな依頼処理: メールを確認しましたが、新規登録対象はありませんでした。"
                send_teams_notification(no_db_msg, config)

            # OneSignal送信
            if total_newly_inserted_main_requests_count > 0:
                send_onesignal_notification(
                    f"{TARGET_CITY} わな依頼 追加",
                    f"{total_newly_inserted_main_requests_count}件の新規依頼がありました。",
                    total_newly_inserted_main_requests_count,
                    config
                )

    except Exception as e_main:
        logging.critical(f"致命的エラー: {e_main}\n{traceback.format_exc()}")

    logging.info("スクリプトを終了します。")

if __name__ == '__main__':
    main()